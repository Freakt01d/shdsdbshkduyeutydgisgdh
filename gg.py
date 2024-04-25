from openpyxl import Workbook
from openpyxl.styles import Border, Side

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Let's assume your pivot table fills from A1 to D10
# First, populate the cells with some example data
for row in range(1, 11):
    for col in range(1, 5):
        ws.cell(row=row, column=col, value=f"Data {row},{col}")

# Define the border style
thin_border = Border(
    left=Side(style='thin'), 
    right=Side(style='thin'), 
    top=Side(style='thin'), 
    bottom=Side(style='thin')
)

# Apply the border to each cell in the range
for row in ws['A1:D10']:
    for cell in row:
        cell.border = thin_border

# Save the workbook
wb.save('styled_pivot_table.xlsx')
