"""
Parallel Trader_IGGID + Sales_IGGID extraction across all auto-discovered
system partitions for redservice.t_raw_detail_audit, 1-17 Apr 2026.

Tuned for client = Standard D4s v3 (4 vCores / 16 GB).

Flow:
  1. Discover partitions ending in _202604.
  2. Phase A (PROBE): run a per-day GROUP BY check on every system in
     parallel. Quick, returns at most 17 rows per system. Reports
     which (system, day) combinations have IGGID and which don't.
  3. Phase B (SCAN): for each system that has at least one qualifying
     day, run the full regex extraction restricted to those days only.
     Up to MAX_WORKERS systems scanned in parallel.
  4. Values stream through a queue to a single writer thread that
     appends to an xlsx in write_only mode (rows go to a temp spool,
     not RAM). Empty Value="" matches are skipped.

Output: trader_sales_iggid_01apr_17apr_2026.xlsx
        Sheet 1: Trader_IGGID
        Sheet 2: Sales_IGGID
"""

import psycopg2
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook

# ----------------------------------------------------------------------
# Connection
# ----------------------------------------------------------------------
CONN = {
    "host":     "redhist-db.postgres.database.azure.com",
    "port":     5432,
    "dbname":   "redservice",
    "user":     "<user>",
    "password": "<password>",
    "sslmode":  "require",
}

# ----------------------------------------------------------------------
# Range
# ----------------------------------------------------------------------
YYYYMM     = "202604"
START_DATE = 20260401
END_DATE   = 20260417

# ----------------------------------------------------------------------
# Tuning (D4s v3 client)
# ----------------------------------------------------------------------
MAX_WORKERS     = 8           # parallel full-scan workers
PROBE_WORKERS   = 14          # probes are cheap, run all at once
HEARTBEAT_EVERY = 10_000      # per-system heartbeat
WRITER_HEARTBEAT = 50_000
QUEUE_MAX       = 100_000     # backpressure buffer (16 GB available)

OUTPUT_XLSX = "trader_sales_iggid_01apr_17apr_2026.xlsx"

REGEX = r'Name="(Trader_IGGID|Sales_IGGID)"\s+Value="([^"]*)"'

# ----------------------------------------------------------------------
# Shared state
# ----------------------------------------------------------------------
write_q = queue.Queue(maxsize=QUEUE_MAX)
SENTINEL = None
print_lock = threading.Lock()

def log(msg):
    with print_lock:
        print(msg, flush=True)


# ----------------------------------------------------------------------
# Partition discovery
# ----------------------------------------------------------------------
def discover_partitions():
    conn = psycopg2.connect(**CONN)
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT n.nspname || '.' || c.relname AS fqtn
                FROM pg_class c
                JOIN pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = 'redservice'
                  AND c.relname LIKE %s
                  AND c.relkind IN ('r', 'p')
                ORDER BY c.relname;
            """, (f"t_raw_detail_audit_%_{YYYYMM}",))
            return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


# ----------------------------------------------------------------------
# Phase A: per-day probe
# ----------------------------------------------------------------------
def probe_system(table):
    sys_label = table.split("_")[-2]
    conn = psycopg2.connect(**CONN)
    try:
        with conn.cursor() as cur:
            cur.execute(f"""
                SELECT audit_date,
                       bool_or(request LIKE '%%_IGGID%%') AS has_iggid
                FROM {table}
                WHERE audit_date BETWEEN %s AND %s
                GROUP BY audit_date
                ORDER BY audit_date;
            """, (START_DATE, END_DATE))
            day_rows = cur.fetchall()

        days_with    = [d for d, has in day_rows if has]
        days_without = [d for d, has in day_rows if not has]
        log(f"[probe:{sys_label}] days_with_iggid={len(days_with)} "
            f"days_without={len(days_without)} total_days_present={len(day_rows)}")
        return sys_label, table, days_with, days_without
    except Exception as e:
        log(f"[probe:{sys_label}] ERROR: {e}")
        return sys_label, table, [], []
    finally:
        conn.close()


# ----------------------------------------------------------------------
# Phase B: full scan on qualifying days
# ----------------------------------------------------------------------
def scan_system(sys_label, table, days_with):
    conn = psycopg2.connect(**CONN)
    try:
        cur = conn.cursor(name=f"iggid_{sys_label}")
        cur.itersize = 5000
        cur.execute(f"""
            SELECT m[1] AS id_type, m[2] AS id_value
            FROM (
                SELECT regexp_matches(request, %s, 'g') AS m
                FROM {table}
                WHERE audit_date = ANY(%s)
                  AND request LIKE '%%_IGGID%%'
            ) x;
        """, (REGEX, days_with))

        n_trader = 0
        n_sales  = 0
        n_total  = 0
        for id_type, id_value in cur:
            n_total += 1
            if not id_value:
                continue
            write_q.put((id_type, id_value))
            if id_type == "Trader_IGGID":
                n_trader += 1
            else:
                n_sales += 1
            if n_total % HEARTBEAT_EVERY == 0:
                log(f"[scan:{sys_label}] {n_total} matches "
                    f"(Trader={n_trader}, Sales={n_sales})")

        cur.close()
        log(f"[scan:{sys_label}] DONE  Trader={n_trader}  Sales={n_sales}")
        return sys_label, n_trader, n_sales
    except Exception as e:
        log(f"[scan:{sys_label}] ERROR: {e}")
        return sys_label, 0, 0
    finally:
        conn.close()


# ----------------------------------------------------------------------
# Writer thread
# ----------------------------------------------------------------------
def writer():
    wb = Workbook(write_only=True)
    ws_trader = wb.create_sheet("Trader_IGGID")
    ws_sales  = wb.create_sheet("Sales_IGGID")
    ws_trader.append(["Trader_IGGID"])
    ws_sales.append(["Sales_IGGID"])

    written = 0
    while True:
        item = write_q.get()
        if item is SENTINEL:
            break
        id_type, id_value = item
        if id_type == "Trader_IGGID":
            ws_trader.append([id_value])
        else:
            ws_sales.append([id_value])
        written += 1
        if written % WRITER_HEARTBEAT == 0:
            log(f"[writer] {written} rows written")

    wb.save(OUTPUT_XLSX)
    log(f"[writer] flushed {written} rows -> {OUTPUT_XLSX}")


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main():
    tables = discover_partitions()
    if not tables:
        log(f"No t_raw_detail_audit_*_{YYYYMM} partitions found.")
        return

    log(f"Discovered {len(tables)} partitions for {YYYYMM}:")
    for t in tables:
        log(f"  {t}")

    # ---------------- Phase A: probe all systems in parallel ----------------
    log("\n--- Phase A: probing all systems (per-day IGGID check) ---")
    probe_results = []
    with ThreadPoolExecutor(max_workers=PROBE_WORKERS) as ex:
        futures = [ex.submit(probe_system, t) for t in tables]
        for fut in as_completed(futures):
            probe_results.append(fut.result())

    to_scan = [(s, t, d) for (s, t, d, _) in probe_results if d]
    skipped = [(s, t) for (s, t, d, _) in probe_results if not d]

    log(f"\n--- Probe summary ---")
    log(f"  systems to scan : {len(to_scan)} -> {[s for s,_,_ in to_scan]}")
    log(f"  systems skipped : {len(skipped)} -> {[s for s,_ in skipped]}")
    if not to_scan:
        log("Nothing to scan. Exiting.")
        return

    # ---------------- Writer up before any scanner pushes -----------------
    writer_thread = threading.Thread(target=writer, daemon=False)
    writer_thread.start()

    # ---------------- Phase B: parallel full scans ------------------------
    log(f"\n--- Phase B: scanning {len(to_scan)} systems with "
        f"{min(MAX_WORKERS, len(to_scan))} workers ---")
    summary = []
    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(to_scan))) as ex:
        futures = [ex.submit(scan_system, s, t, d) for s, t, d in to_scan]
        for fut in as_completed(futures):
            summary.append(fut.result())

    write_q.put(SENTINEL)
    writer_thread.join()
