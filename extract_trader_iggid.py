"""
Parallel Trader_IGGID + Sales_IGGID extraction across all auto-discovered
system partitions for redservice.t_raw_detail_audit, 20-24 Apr 2026.

Behavior:
  - Auto-discovers partitions ending in _202604 (skips systems that don't have one).
  - One thread per system, all run in parallel.
  - PROBE: each thread first scans the first 100,000 rows in its partition.
           If none of them contain Trader_IGGID or Sales_IGGID at all,
           the system is skipped entirely (no full scan).
  - Values are streamed through a queue to a single writer thread that
    appends them to an xlsx in write_only mode (rows are flushed to a
    temp file as they arrive, not buffered in memory until the end).
  - Empty Value="" matches are skipped.

Tuned for an 8GB / 2-core client.

Output: trader_sales_iggid_20apr_24apr_2026.xlsx
        Sheet 1: Trader_IGGID
        Sheet 2: Sales_IGGID
"""

import psycopg2
import threading
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook

# ----------------------------------------------------------------------
# Connection
# ----------------------------------------------------------------------
CONN = {
    "host":     "redhist-db.postgres.database.azure.com",
    "port":     5432,
    "dbname":   "redservice",
    "user":     "<user>",
    "password": "<password>",
    "sslmode":  "require",
}

# ----------------------------------------------------------------------
# Range
# ----------------------------------------------------------------------
YYYYMM     = "202604"
START_DATE = 20260420
END_DATE   = 20260424

PROBE_LIMIT     = 100_000     # rows to probe before deciding to skip
HEARTBEAT_EVERY = 10_000      # log per-system progress every N matches
QUEUE_MAX       = 20_000      # backpressure for slow writer

OUTPUT_XLSX = "trader_sales_iggid_20apr_24apr_2026.xlsx"

REGEX = r'Name="(Trader_IGGID|Sales_IGGID)"\s+Value="([^"]*)"'

# ----------------------------------------------------------------------
# Shared state
# ----------------------------------------------------------------------
write_q = queue.Queue(maxsize=QUEUE_MAX)
SENTINEL = None
print_lock = threading.Lock()

def log(msg):
    with print_lock:
        print(msg, flush=True)


# ----------------------------------------------------------------------
# Partition discovery
# ----------------------------------------------------------------------
def discover_partitions():
    conn = psycopg2.connect(**CONN)
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT n.nspname || '.' || c.relname AS fqtn
                FROM pg_class c
                JOIN pg_namespace n ON n.oid = c.relnamespace
                WHERE n.nspname = 'redservice'
                  AND c.relname LIKE %s
                  AND c.relkind IN ('r', 'p')
                ORDER BY c.relname;
            """, (f"t_raw_detail_audit_%_{YYYYMM}",))
            return [r[0] for r in cur.fetchall()]
    finally:
        conn.close()


# ----------------------------------------------------------------------
# Per-system worker
# ----------------------------------------------------------------------
def worker(table):
    sys_label = table.split("_")[-2]   # e.g. astro
    conn = psycopg2.connect(**CONN)
    try:
        # ---- Probe: first PROBE_LIMIT rows in date range ----
        with conn.cursor() as probe:
            probe.execute(f"""
                SELECT COALESCE(bool_or(request LIKE '%%_IGGID%%'), false)
                FROM (
                    SELECT request
                    FROM {table}
                    WHERE audit_date BETWEEN %s AND %s
                    LIMIT %s
                ) x;
            """, (START_DATE, END_DATE, PROBE_LIMIT))
            has_iggid = probe.fetchone()[0]

        if not has_iggid:
            log(f"[{sys_label}] no IGGID found in first {PROBE_LIMIT} rows -- skipping")
            return sys_label, 0, 0, True   # skipped=True

        log(f"[{sys_label}] probe positive, starting full scan")

        # ---- Full scan with server-side cursor ----
        cur = conn.cursor(name=f"iggid_{sys_label}")
        cur.itersize = 5000
        cur.execute(f"""
            SELECT m[1] AS id_type, m[2] AS id_value
            FROM (
                SELECT regexp_matches(request, %s, 'g') AS m
                FROM {table}
                WHERE audit_date BETWEEN %s AND %s
                  AND request LIKE '%%_IGGID%%'
            ) x;
        """, (REGEX, START_DATE, END_DATE))

        n_trader = 0
        n_sales  = 0
        n_total  = 0
        for id_type, id_value in cur:
            n_total += 1
            if not id_value:
                continue
            write_q.put((id_type, id_value))
            if id_type == "Trader_IGGID":
                n_trader += 1
            else:
                n_sales += 1
            if n_total % HEARTBEAT_EVERY == 0:
                log(f"[{sys_label}] {n_total} matches so far "
                    f"(Trader={n_trader}, Sales={n_sales})")

        cur.close()
        log(f"[{sys_label}] DONE  Trader={n_trader}  Sales={n_sales}")
        return sys_label, n_trader, n_sales, False
    except Exception as e:
        log(f"[{sys_label}] ERROR: {e}")
        return sys_label, 0, 0, False
    finally:
        conn.close()


# ----------------------------------------------------------------------
# Writer thread (single, owns the workbook)
# ----------------------------------------------------------------------
def writer():
    wb = Workbook(write_only=True)
    ws_trader = wb.create_sheet("Trader_IGGID")
    ws_sales  = wb.create_sheet("Sales_IGGID")
    ws_trader.append(["Trader_IGGID"])
    ws_sales.append(["Sales_IGGID"])

    written = 0
    while True:
        item = write_q.get()
        if item is SENTINEL:
            break
        id_type, id_value = item
        if id_type == "Trader_IGGID":
            ws_trader.append([id_value])
        else:
            ws_sales.append([id_value])
        written += 1
        if written % 50_000 == 0:
            log(f"[writer] {written} rows written to xlsx")

    wb.save(OUTPUT_XLSX)
    log(f"[writer] flushed {written} rows -> {OUTPUT_XLSX}")


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main():
    tables = discover_partitions()
    if not tables:
        log(f"No t_raw_detail_audit_*_{YYYYMM} partitions found.")
        return

    log(f"Discovered {len(tables)} partitions for {YYYYMM}:")
    for t in tables:
        log(f"  {t}")

    writer_thread = threading.Thread(target=writer, daemon=False)
    writer_thread.start()

    summary = []
    with ThreadPoolExecutor(max_workers=len(tables)) as ex:
        futures = {ex.submit(worker, t): t for t in tables}
        for fut in as_completed(futures):
            summary.append(fut.result())

    # tell writer to finish
    write_q.put(SENTINEL)
    writer_thread.join()

    # final report
    log("\n========= SUMMARY =========")
    total_t = total_s = 0
    for sys_label, nt, ns, skipped in sorted(summary):
        tag = "SKIPPED" if skipped else "scanned"
        log(f"  {sys_label:12s} {tag}  Trader={nt:>8d}  Sales={ns:>8d}")
        total_t += nt
        total_s += ns
    log(f"  {'TOTAL':12s}          Trader={total_t:>8d}  Sales={total_s:>8d}")
    log(f"Saved -> {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
